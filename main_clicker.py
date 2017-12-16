import time
import openpyxl
import error_hook
from selenium.webdriver.firefox.webdriver import FirefoxProfile
from freedogeco import FreeDOGECO
from multicoin import MulCo


print("\n\n==Добро пожаловать в автоматизированную систему заработка на кранах MoneyCrane==")
print("\t\t\t\t\t\t\t==Автор: \n\n\n")
mulco_login = ""
dogeco_mail = ""

# ___________________________________MULTICOINFAUCET.COM / INFORMATION____________________________________________
def init_mulco():
    try:
        info_wb = openpyxl.load_workbook(filename=r'data\crane_stats.xlsx')
        info_sheet = info_wb['INFO']
    except Exception as err:
        error_hook.fatal_err(
            "Не могу получить доступ к crane_stats.xlsx/INFO. "
            "Убедитесь в наличии файла 'crane_stats.xlsx' в папке 'data' и повторите попытку."
            "\n Файл должен быть закрыт!", err)
        return None
    if not info_sheet.cell(row=10, column=10).value:
        error_hook.warn_err("В ячейке 'J10' не указан пароль. Работа с сайтом multicoinfaucet.com невозможна")
        return None
    else:
        mulco_pass = str(info_sheet.cell(row=10, column=10).value)
    mulco_user = MulCo(mulco_login, mulco_pass)
    return mulco_user
# ________________________________________________________________________________________________________________


# ____________________________________________DOGE.CO.IN / INFORMATION____________________________________________
def init_doge():
    try:
        info_wb = openpyxl.load_workbook(filename=r'data\crane_stats.xlsx')
        info_sheet = info_wb['INFO']
    except Exception as err:
        error_hook.fatal_err(
            "Не могу получить доступ к crane_stats.xlsx/INFO. "
            "Убедитесь в наличии файла 'crane_stats.xlsx' в папке 'data' и повторите попытку."
            "\n Файл должен быть закрыт!", err)
        return None
    dogeco_cap_img = 'dgccap.png'
    if not info_sheet.cell(row=10, column=2).value:
        error_hook.warn_err("В ячейке 'B10' не указан пароль. Работа с сайтом fredoge.co.in невозможна")
        return None
    else:
        dogeco_pass = str(info_sheet.cell(row=10, column=2).value)
    if not info_sheet.cell(row=19, column=9).value:
        error_hook.warn_err("В ячейке 'i19' не указан ключ каптчи. Работа с сайтом fredoge.co.in невозможна")
        return None
    else:
        dogeco_cap_api = str(info_sheet.cell(row=19, column=9).value)
    if not info_sheet.cell(row=20, column=9).value:
        error_hook.warn_err("В ячейке 'i20' не указан профиль. Работа с сайтом fredoge.co.in невозможна")
        return None
    else:
        profile_dgc = FirefoxProfile(str(info_sheet.cell(row=20, column=9).value))
    dogeco_user = FreeDOGECO(dogeco_mail, dogeco_pass, profile_dgc, dogeco_cap_img, dogeco_cap_api)
    return dogeco_user
# ________________________________________________________________________________________________________________

try:
    fdgc_sheet = openpyxl.load_workbook(filename=r'data\crane_stats.xlsx')
    fdgc_wb = fdgc_sheet['INFO']
    fdgc_wb['D5'] = ''
    fdgc_wb['L5'] = ''
except Exception as err:
    error_hook.fatal_err(
        "Не могу записать время начала пула. "
        "Убедитесь в наличии файла 'crane_stats.xlsx' в папке 'data' и повторите попытку."
        "\n Файл должен быть закрыт!", err)
    raise SystemExit(404)

while 1:
    a = time.strftime("%H:%M:%S", time.localtime())
    print("\n\n_____[+]|" + str(a) + "|Начинаю работу с сайтом freedoge.co.in_____")
    dogeco_user = init_doge()
    dgc_cooldown = 3600
    if dogeco_user:
        dgc_cooldown = FreeDOGECO.roller(dogeco_user)
        a = time.strftime("[%H:%M:%S]", time.localtime())
        if not dgc_cooldown:
            a = time.strftime("%H:%M:%S", time.localtime())
            error_hook.fatal_err("В процессе работы произошла непредвиденная ошибка. Повторю попытку через 10 минут", "UNKNOWN")
            dgc_cooldown = 600
    print("_____[+]|" + str(a) + "|Работа с сайтом freedoge.co.in завершена_____")
    #
    #
    #
    a = time.strftime("%H:%M:%S", time.localtime())
    print("\n\n_____[+]|" + str(a) + "|Начинаю работу с сайтом multicoinfaucet.com_____")
    mulco_user = init_mulco()
    mulco_cooldown = 3600
    if mulco_user:
        mulco_cooldown = MulCo.start(mulco_user)
        a = time.strftime("[%H.%M.%S]", time.localtime())
        if not mulco_cooldown:
            a = time.strftime("%H:%M:%S", time.localtime())
            error_hook.fatal_err("В процессе работы произошла непредвиденная ошибка. Повторю попытку через 10 минут", "UNKNOWN")
            mulco_cooldown = 600
    print("_____[+]|" + str(a) + "|Работа с сайтом multicoinfaucet.com завершена_____\n")
    #
    #
    wait_secs = mulco_cooldown
    if int(dgc_cooldown) < int(mulco_cooldown):
            wait_secs = dgc_cooldown
    print("\n\n[+]|" + str(a) + "|Цикл завершен, жду " + str(wait_secs) + " сек...\n")
    time.sleep(wait_secs)




