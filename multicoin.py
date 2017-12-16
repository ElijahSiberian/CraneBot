import requests
import time
import re
from datetime import datetime       #to measure time
import csv                          #to write in csv
import error_hook
from selenium import webdriver
import openpyxl


class MulCo:
    btc_crane_url = "https://multicoinfaucet.com/btc-faucet/roll"
    permanent_balance = 0.0
    mail = ' '
    passw = ' '
    def __init__(self, name, passw):
        self.mail = str(name)
        self.passw = str(passw)
        print("\nИнициализация аккаунта на multicoinfaucet.com..." +
              "\n\t[#]_LOGIN_= " + self.mail +
              "\n\t[#]_PASSW_= " + self.passw)


    def start(self):
        a = time.strftime("%H:%M:%S", time.localtime())
        try:
            fdgc_wb = openpyxl.load_workbook(filename=r'data\crane_stats.xlsx')
            fdgc_sheet = fdgc_wb['INFO']
            fdgc_sheet['L15'] = str(a)
            fdgc_wb.save(r'data\crane_stats.xlsx')
        except Exception as err:
            error_hook.fatal_err(
                "Не могу получить доступ к файлу 'crane_stats.xlsx'. "
                "Убедитесь в наличии файла 'crane_stats.xlsx' в папке 'data' и повторите попытку."
                "\n Файл должен быть закрыт!", err)
        try:
            browser = webdriver.Firefox()
            browser.get("https://multicoinfaucet.com")
        except Exception as err:
            error_hook.fatal_err("Не могу зайти на сайт", err)
            return None
        time.sleep(2)
        try:
            login_but = browser.find_element_by_xpath('/html/body/section[1]/div/div[3]/a[1]')
            login_but.click()
        except Exception as err:
            error_hook.fatal_err("Не могу найти объект", err)
            return None
        time.sleep(2)
        for i in range(1, 100):
            try:
                login_inp = browser.find_element_by_xpath('//*[@id="username"]')
                login_inp.send_keys(self.mail)
                pass_inp = browser.find_element_by_xpath('//*[@id="password"]')
                pass_inp.send_keys(self.passw)
            except Exception as err:
                error_hook.fatal_err("Не могу найти объект", err)
                return None
            try:
                guess = browser.find_element_by_xpath('//*[@id="visualCaptcha-img-0"]')
                print("[.]Пытаюсь угадать капчу...")
                guess.click()
            except Exception as err:
                error_hook.fatal_err("Не удалось найти картинку капчи", err)
                error_hook.screen_grab(browser)
            try:
                login_but = browser.find_element_by_xpath('/html/body/div[2]'
                                                          '/section/div/div/div'
                                                          '/form/input[5]')
                login_but.click()
                time.sleep(2)
                goto_btc = browser.find_element_by_xpath('/html/body/section[1]/div/div[1]/a')
                goto_btc.click()
                break
            except:
                print("[.]Не угадал, попробую еще раз")
                continue
        try:
            goto_roll = browser.find_element_by_xpath('/html/body/div[3]/div/div[2]/div[1]/div/div/a')
            goto_roll.click()
        except Exception as err:
            error_hook.fatal_err("Не могу найти объект", err)
            return None
        time.sleep(2)
        print('[+]Я зашел на сайт')
        return MulCo.roller(self, browser)

    def roller(self, browser):
        start_mulco = datetime.now()
        try:
            balance = browser.find_element_by_xpath('//*[@id="balance_global"]')
            balance = balance.text
        except:
            error_hook.warn_err("Не удалось узнать баланс")
            error_hook.screen_grab(browser)
        else:
            print("[+]Ваш текущий баланс = " + balance + " BTC")
            self.permanent_balance = float(balance)
            try:
                fdgc_wb = openpyxl.load_workbook(filename=r'data\crane_stats.xlsx')
                fdgc_sheet = fdgc_wb['INFO']
                fdgc_sheet.cell(row=4, column=12).value = "%.8f" % float(balance)
                fdgc_wb.save(r'data\crane_stats.xlsx')
            except Exception as err:
                error_hook.fatal_err("Не могу получить доступ к crane_stats.xlsx/INFO", err)

        try:
            time_to_wait = browser.find_element_by_xpath('/html/body/section'
                                                         '/div/div/div/div/div'
                                                         '/div/div[12]/center/span[2]/b')
        except Exception:
            print('[.]Ожидание не требуется - приступаю к работе')
        else:
            try:
                minute = re.split(' ', time_to_wait.text)
                print("[.]Нужно подождать " + minute[5] + " min")
                secs_to_wait = int(minute[5]) * 60
                browser.close()
                return secs_to_wait
            except Exception as err:
                browser.close()
                error_hook.fatal_err("", err)
                return None

        for i in range(1, 100):
            time.sleep(1)
            try:
                guess = browser.find_element_by_xpath('//*[@id="visualCaptcha-img-0"]')
                print("[.]Пытаюсь угадать капчу...")
                guess.click()
            except Exception as err:
                error_hook.warn_err("Не удалось найти картинку капчи")
                try:
                    resp = browser.find_element_by_xpath('//*[@id="responseMessage"]')
                    print("[+]Пул окончен")
                    print("[%]\n" + resp.text + "\n[%]")
                    break
                except:
                    error_hook.fatal_err("Не удалось сделать пул", err)
                    error_hook.screen_grab(browser)
                    browser.close()
                    return None
            try:
                button = browser.find_element_by_xpath('/html/body/section'
                                                       '/div/div/div/div/div/div'
                                                       '/div[12]/center/form/div/input')
                button.click()
                time.sleep(2)
                resp = browser.find_element_by_xpath('//*[@id="responseMessage"]')
                trouble_cap = re.match('CAPTCHA failed: No CAPTCHA data recieved '
                                       'in POST..Please retry', resp.text)
                if trouble_cap:
                    print("[.]Обновляю POST-запрос...")
                    browser.get(browser.current_url)
            except Exception as err:
                error_hook.warn_err("Не могу найти кнопку ROLL")
            else:
                print("[.]Не угадал, попробую еще раз")
                continue
        try:
            balance = browser.find_element_by_xpath('//*[@id="balance_global"]')
            balance = balance.text
        except:
            error_hook.warn_err("Не удалось узнать баланс")
            error_hook.screen_grab(browser)
        else:
            print("[+]Ваш текущий баланс = " + balance + " BTC")
            income = float(balance) - float(self.permanent_balance)
            try:
                fdgc_wb = openpyxl.load_workbook(filename=r'data\crane_stats.xlsx')
                fdgc_sheet = fdgc_wb['INFO']
                fdgc_sheet.cell(row=4, column=12).value = "%.8f" % float(balance)
                fdgc_wb.save(r'data\crane_stats.xlsx')
            except Exception as err:
                error_hook.fatal_err("Не могу получить доступ к crane_stats.xlsx/INFO", err)
            print("[+]Ваша прибыль +%.8f" % float(income) + " BTC")
            self.permanent_balance = float(balance)
            try:
                fdgc_wb = openpyxl.load_workbook(filename=r'data\crane_stats.xlsx')
                fdgc_sheet = fdgc_wb['INFO']
                fdgc_sheet.cell(row=5, column=12).value = "%.8f" % float(income)
                if float(fdgc_sheet.cell(row=5, column=12).value) > float(fdgc_sheet.cell(row=6, column=12).value):
                    fdgc_sheet.cell(row=6, column=12).value = "%.8f" % float(income)
                fdgc_wb.save(r'data\crane_stats.xlsx')
            except Exception as err:
                error_hook.fatal_err("Не могу получить доступ к crane_stats.xlsx/INFO", err)
        browser.close()
        end_mulco = datetime.now()
        total_mulco = end_mulco - start_mulco
        total_mulco = int(total_mulco.total_seconds())
        print("[+]Цикл занял " + str(total_mulco) + " sec")
        return 3600