import requests
import time
import re as reg_exp
from datetime import datetime       #to measure time
import csv                          #to write in csv
import error_hook
from rucaptcha import RuCaptcha
from selenium import webdriver
import openpyxl
import selenium.common.exceptions as sel_error


class FreeDOGECO:
    cap_api = ' '
    cap_img = ' '
    mail = ' '
    passw = ' '
    captcha_to_use = "solvemedia"
    CAP = RuCaptcha()

    def __init__(self, mail, passw, profile, cap_img, cap_api):
        self.mail = mail
        self.passw = passw
        self.profile = profile
        self.cap_img = cap_img
        self.cap_api = cap_api
        print("\nИнициализация аккаунта на freedoge.co.in..." +
              "\n\t[#]_LOGIN_= " + self.mail +
              "\n\t[#]_PASSW_= " + self.passw +
              "\n\t[#]PROFILE= " + str(self.profile.profile_dir) +
              "\n\t[#]CAP_IMG= " + self.cap_img +
              "\n\t[#]API_KEY= " + self.cap_api)

    def roller(self):
        a = time.strftime("%H:%M:%S", time.localtime())
        try:
            fdgc_wb = openpyxl.load_workbook(filename=r'data\crane_stats.xlsx')
            fdgc_sheet = fdgc_wb['INFO']
            fdgc_sheet['D15'] = str(a)
            fdgc_wb.save(r'data\crane_stats.xlsx')
        except Exception as err:
            error_hook.fatal_err(
                "Не могу записать время начала пула. "
                "Убедитесь в наличии файла 'crane_stats.xlsx' в папке 'data' и повторите попытку."
                "\n Файл должен быть закрыт!", err)
        try:
            browser = webdriver.Firefox(firefox_profile=self.profile)
            browser.get("https://freedoge.co.in/?op=home")
        except Exception as err:
            error_hook.fatal_err("Не могу зайти на сайт", err)
            return None
        start_rollFDCO = datetime.now()
        time.sleep(3)
        bal_1 = 0.0
        bal_2 = 0.0
        try:
            balance = browser.find_element_by_xpath("//*[@id='balance']")
            bal_1 = float(balance.text)
            print("[+]Текущий баланс " + str(bal_1) + " DOGE")
            try:
                fdgc_wb = openpyxl.load_workbook(filename=r'data\crane_stats.xlsx')
                fdgc_sheet = fdgc_wb['INFO']
                fdgc_sheet.cell(row=4, column=4).value = "%.8f" % bal_1
                fdgc_wb.save(r'data\crane_stats.xlsx')
            except Exception as err:
                error_hook.fatal_err("Не могу получить доступ к crane_stats.xlsx/INFO", err)
        except Exception:
            error_hook.warn_err("Не удалось узнать баланс")
        time.sleep(5)
        try:
            element = browser.find_element_by_xpath('//*[@id="free_play_captcha_types"]')
            all_options = element.find_elements_by_tag_name("option")
            print("[+]Выбор из: ")
            for option in all_options:
                print("\t - %s" % option.get_attribute("value"))
                if option.get_attribute("value") == self.captcha_to_use:
                    print("[+]Выбрал " + self.captcha_to_use)
                    option.click()
        except Exception:
            error_hook.warn_err("Не удалось выбрать капчу")
            return FreeDOGECO.check_timer(self, browser)
        time.sleep(8)
        solved_cap = RuCaptcha.captcha_solver(self.CAP, self.cap_img, self.cap_api, browser)
        if not solved_cap:
            error_hook.warn_err("Не удалось решить капчу")
            browser.close()
            return None
        try:
            captcha_point = browser.find_element_by_id("adcopy_response")
            captcha_point.clear()
            captcha_point.send_keys(str(solved_cap))
        except Exception as err:
            error_hook.fatal_err("Не удалось ввести капчу",err)
            browser.close()
            return FreeDOGECO.roller(self)
        try:
            # rollbutton = browser.find_element_by_xpath("/html/body/div[3]/div[1]/div[1]/div[2]/p[2]/input[1]")
            rollbutton = browser.find_element_by_xpath('//*[@id="free_play_form_button"]')
            browser.execute_script("return arguments[0].scrollIntoView();", rollbutton)
            rollbutton.click()
        except Exception:
            error_hook.warn_err("Не могу найти кнопку [ROLL!].Попробую еще раз...")
            return FreeDOGECO.check_timer(self, browser)
        print("[+]Я заролил!")
        time.sleep(2)
        try:
            closebutton = browser.find_element_by_xpath("/html/body/div[11]/a[1]")
            closebutton.click()
        except Exception:
            error_hook.warn_err("Не удалось закрыть всплывающее окно")
        try:
            browser.refresh()
            balance = browser.find_element_by_xpath("//*[@id='balance']")
            bal_2 = float(balance.text)
            print("[+]Баланс  =  " + str(bal_2) + " DOGE")
            try:
                fdgc_wb = openpyxl.load_workbook(filename=r'data\crane_stats.xlsx')
                fdgc_sheet = fdgc_wb['INFO']
                fdgc_sheet.cell(row=4, column=4).value = "%.8f" % float(bal_2)
                fdgc_wb.save(r'data\crane_stats.xlsx')
            except Exception as err:
                error_hook.fatal_err("Не могу получить доступ к crane_stats.xlsx/INFO", err)
            winnings = float(bal_2) - float(bal_1)
            print("[+]Прибыль = +" + ("%.2f" % winnings) + " DOGE")
            try:
                fdgc_wb = openpyxl.load_workbook(filename=r'data\crane_stats.xlsx')
                fdgc_sheet = fdgc_wb['INFO']
                fdgc_sheet.cell(row=5, column=4).value = "%.8f" % float(winnings)
                if float(fdgc_sheet.cell(row=5, column=4).value) > float(fdgc_sheet.cell(row=6, column=4).value):
                    fdgc_sheet.cell(row=6, column=4).value = "%.8f" % float(winnings)
                fdgc_wb.save(r'data\crane_stats.xlsx')
            except Exception as err:
                error_hook.fatal_err("Не могу получить доступ к crane_stats.xlsx/INFO", err)
        except Exception:
            error_hook.warn_err("Не удалось узнать баланс")


        end_rollFDCO = datetime.now() - start_rollFDCO
        browser.close()
        print("[%]Время, затраченное на ролл: " + str(end_rollFDCO))
        print("[.]Ожидаю следующего ролла...\n\n")
        return 3600

    def check_timer(self, browser):
        try:
            split_title = reg_exp.split(r'm', browser.title, maxsplit=1)
            browser.close()
            time_clock = int(split_title[0]) + 1
            fckng_wrd_dmn_it = ''
            if (time_clock % 10) == 1:
                fckng_wrd_dmn_it = ' минуту'
            elif time_clock < 5:
                fckng_wrd_dmn_it = ' минуты'
            else:
                fckng_wrd_dmn_it = ' минут'
            print("[+]Кажется, мне стоит попробовать еще раз через " + str(time_clock) + fckng_wrd_dmn_it)
            time_to_wait = time_clock * 60
            return time_to_wait
        except Exception as err:
            error_hook.fatal_err("Я до сих пор не могу найти кнопку [ROLL!]. Кажется, что-то пошло не так!", err)
            return None
